import boto3
import argparse
import logging
import datetime
import openpyxl

def setup_logging():
    logging.basicConfig(
        format='%(asctime)s [%(levelname)s] %(message)s',
        level=logging.INFO,
        datefmt='%Y-%m-%d %H:%M:%S'
    )

def initialize_aws_session(profile, region):
    try:
        return boto3.Session(profile_name=profile, region_name=region)
    except Exception as e:
        logging.error(f"Failed to initialize AWS session: {e}")
        return None

def get_ec2_costs(session, start_date, end_date):
    try:
        cost_explorer = session.client('ce')
        response = cost_explorer.get_cost_and_usage(
            TimePeriod={'Start': start_date, 'End': end_date},
            Granularity='MONTHLY',
            Metrics=['NetAmortizedCost'],
            GroupBy=[{'Type': 'DIMENSION', 'Key': 'INSTANCE_TYPE'}]
        )
        return {item['Keys'][0]: float(item['Metrics']['NetAmortizedCost']['Amount'])
                for item in response['ResultsByTime'][0]['Groups']}
    except Exception as e:
        logging.error(f"Failed to retrieve EC2 costs: {e}")
        return {}

def get_ebs_costs(session, start_date, end_date):
    try:
        cost_explorer = session.client('ce')
        response = cost_explorer.get_cost_and_usage(
            TimePeriod={'Start': start_date, 'End': end_date},
            Granularity='MONTHLY',
            Metrics=['NetAmortizedCost'],
            GroupBy=[{'Type': 'DIMENSION', 'Key': 'USAGE_TYPE'}]
        )
        return {item['Keys'][0]: float(item['Metrics']['NetAmortizedCost']['Amount'])
                for item in response['ResultsByTime'][0]['Groups']}
    except Exception as e:
        logging.error(f"Failed to retrieve EBS costs: {e}")
        return {}

def get_snapshot_costs(session, start_date, end_date):
    try:
        cost_explorer = session.client('ce')
        response = cost_explorer.get_cost_and_usage(
            TimePeriod={'Start': start_date, 'End': end_date},
            Granularity='MONTHLY',
            Metrics=['NetAmortizedCost'],
            GroupBy=[{'Type': 'DIMENSION', 'Key': 'USAGE_TYPE'}]
        )
        return {item['Keys'][0]: float(item['Metrics']['NetAmortizedCost']['Amount'])
                for item in response['ResultsByTime'][0]['Groups'] if "EBS:SnapshotUsage" in item['Keys'][0]}
    except Exception as e:
        logging.error(f"Failed to retrieve snapshot costs: {e}")
        return {}

def get_ec2_instances(session, ec2_cost_mapping):
    ec2_client = session.client('ec2')
    instances = []
    try:
        response = ec2_client.describe_instances()
        for reservation in response['Reservations']:
            for instance in reservation['Instances']:
                instance_id = instance['InstanceId']
                state = instance['State']['Name']
                instance_type = instance['InstanceType']
                ec2_cost = ec2_cost_mapping.get(instance_type, 0.0)
                name_tag = next((tag['Value'] for tag in instance.get('Tags', []) if tag['Key'] == 'Name'), 'N/A')

                instances.append([name_tag, instance_id, state, ec2_cost])
        return instances
    except Exception as e:
        logging.error(f"Failed to retrieve EC2 instances: {e}")
        return []

def get_ebs_volumes(session, ebs_cost_mapping):
    ebs_client = session.client('ec2')
    volumes = []
    try:
        response = ebs_client.describe_volumes()
        for volume in response['Volumes']:
            volume_id = volume['VolumeId']
            volume_type = volume['VolumeType']
            volume_cost = ebs_cost_mapping.get(volume_type, 0.0)
            name_tag = next((tag['Value'] for tag in volume.get('Tags', []) if tag['Key'] == 'Name'), 'N/A')

            volumes.append([volume_id, name_tag, volume_cost])
        return volumes
    except Exception as e:
        logging.error(f"Failed to retrieve EBS volumes: {e}")
        return []

def get_snapshots(session, snapshot_cost_mapping):
    ec2_client = session.client('ec2')
    snapshots = []
    try:
        response = ec2_client.describe_snapshots(OwnerIds=['self'])
        for snapshot in response['Snapshots']:
            snapshot_id = snapshot['SnapshotId']
            volume_id = snapshot.get('VolumeId', 'N/A')
            size_gb = snapshot['VolumeSize']
            description = snapshot.get('Description', 'N/A')
            status = snapshot['State']
            encryption = snapshot['Encrypted']
            kms_key_id = snapshot.get('KmsKeyId', 'N/A')

            snapshot_cost = snapshot_cost_mapping.get("EBS:SnapshotUsage", 0.0)

            snapshots.append([snapshot_id, volume_id, snapshot_cost, size_gb, description, status, encryption, kms_key_id])
        return snapshots
    except Exception as e:
        logging.error(f"Failed to retrieve snapshots: {e}")
        return []

def export_to_excel(ec2_data, ebs_data, snapshot_data, output_file):
    try:
        wb = openpyxl.Workbook()
        
        ws1 = wb.active
        ws1.title = "EC2_Cost_Report"
        ws1.append(["EC2 Instance Name", "EC2 Instance ID", "EC2 Instance State", "EC2 Instance Cost"])
        for row in ec2_data:
            ws1.append(row)

        ws2 = wb.create_sheet(title="EBS_Cost_Report")
        ws2.append(["EBS Volume ID", "EBS Volume Name", "EBS Volume Cost"])
        for row in ebs_data:
            ws2.append(row)

        ws3 = wb.create_sheet(title="Snapshots_Cost_Report")
        ws3.append(["Snapshot ID", "Volume ID", "Snapshot Cost", "Full Snapshot Size", "Description", "Snapshot Status", "Encryption", "KMS Key ID"])
        for row in snapshot_data:
            ws3.append(row)

        wb.save(output_file)
        logging.info(f"Report successfully saved to {output_file}")
    except Exception as e:
        logging.error(f"Failed to export to Excel: {e}")

def main():
    setup_logging()
    parser = argparse.ArgumentParser(description="AWS Cost Report Generator")
    parser.add_argument("--profile", required=True, help="AWS profile name")
    parser.add_argument("--region", required=True, help="AWS region")
    args = parser.parse_args()

    session = initialize_aws_session(args.profile, args.region)
    if not session:
        return

    start_date = "2025-02-01"
    end_date = "2025-02-28"

    ec2_cost_mapping = get_ec2_costs(session, start_date, end_date)
    ebs_cost_mapping = get_ebs_costs(session, start_date, end_date)
    snapshot_cost_mapping = get_snapshot_costs(session, start_date, end_date)

    ec2_data = get_ec2_instances(session, ec2_cost_mapping)
    ebs_data = get_ebs_volumes(session, ebs_cost_mapping)
    snapshot_data = get_snapshots(session, snapshot_cost_mapping)

    export_to_excel(ec2_data, ebs_data, snapshot_data, "aws_cost_report.xlsx")

if __name__ == "__main__":
    main()
